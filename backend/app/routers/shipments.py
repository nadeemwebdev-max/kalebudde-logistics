import csv
import io
import os
import random
import string
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response, StreamingResponse
import openpyxl
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import get_current_user, require_admin, require_staff
from app.models import Role, Shipment, ShipmentStatus, TrackingEvent, User
from app.schemas import (
    PublicTrackingOut,
    ShipmentCreate,
    ShipmentOut,
    ShipmentUpdate,
    TrackingEventCreate,
    TrackingEventOut,
)

router = APIRouter(prefix="/api", tags=["shipments"])

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.post("/shipments/upload-lr")
def upload_lr_copy(
    file: UploadFile = File(...),
    user: User = Depends(require_staff),
):
    """Upload LR Copy document (PDF, PNG, JPG) and return public URL."""
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in [".pdf", ".png", ".jpg", ".jpeg", ".webp"]:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file format. Please upload PDF, PNG, JPG, or WEBP.",
        )

    filename = f"LR_{uuid.uuid4().hex[:12]}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)

    with open(filepath, "wb") as f:
        f.write(file.file.read())

    return {"lr_copy_url": f"/uploads/{filename}", "filename": filename}


def generate_tracking_number(db: Session) -> str:
    while True:
        code = "KL" + "".join(random.choices(string.digits, k=9))
        if not db.query(Shipment).filter(Shipment.tracking_number == code).first():
            return code


def parse_date_str(val: str | None) -> datetime | None:
    if not val or not str(val).strip():
        return None
    val_str = str(val).strip()
    formats = [
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(val_str, fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


# ---------- public tracker ----------
@router.get("/track/{tracking_number}", response_model=PublicTrackingOut)
def public_track(tracking_number: str, db: Session = Depends(get_db)):
    code = tracking_number.strip()
    s = (
        db.query(Shipment)
        .filter(
            (Shipment.tracking_number.ilike(code))
            | (Shipment.lr_number.ilike(code))
            | (Shipment.invoice_number.ilike(code))
            | (Shipment.eway_bill_number.ilike(code))
        )
        .first()
    )
    if not s:
        raise HTTPException(status_code=404, detail="Tracking number or LR number not found")
    return s


# ---------- authenticated list & single ----------
@router.get("/shipments", response_model=list[ShipmentOut])
def list_shipments(
    status_filter: ShipmentStatus | None = Query(None, alias="status"),
    q: str | None = None,
    limit: int = Query(200, le=500),
    offset: int = 0,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    query = db.query(Shipment)
    if user.role == Role.CLIENT:
        query = query.filter(Shipment.client_id == user.id)
    if status_filter:
        query = query.filter(Shipment.status == status_filter)
    if q:
        like = f"%{q}%"
        query = query.filter(
            Shipment.tracking_number.ilike(like)
            | Shipment.lr_number.ilike(like)
            | Shipment.invoice_number.ilike(like)
            | Shipment.eway_bill_number.ilike(like)
            | Shipment.consignor.ilike(like)
            | Shipment.consignee.ilike(like)
            | Shipment.origin.ilike(like)
            | Shipment.destination.ilike(like)
            | Shipment.driver_name.ilike(like)
        )
    return query.order_by(Shipment.created_at.desc()).offset(offset).limit(limit).all()


# ---------- EXCEL & CSV DATA EXPORT & SAMPLE TEMPLATE ----------
HEADERS = [
    "LR NUMBER",
    "INVOICE",
    "INVOICE DATE",
    "E-WAY BILL NUMBER",
    "E-WAY BILL DATE",
    "E-WAY BILL EXPIRY DATE",
    "E-WAY BILL STATUS",
    "NEW EXTENDED E-WAY BILL DATE",
    "ORIGIN",
    "DESTINATION",
    "CONSIGNOR",
    "CONSIGNEE",
    "DRIVER NAME",
    "DRIVER NO",
    "DELIVERY STATUS",
]


@router.get("/shipments/export-csv")
def export_csv_shipments(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    query = db.query(Shipment)
    if user.role == Role.CLIENT:
        query = query.filter(Shipment.client_id == user.id)
    shipments = query.order_by(Shipment.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)

    for s in shipments:
        writer.writerow([
            s.lr_number or s.tracking_number,
            s.invoice_number or "",
            s.invoice_date.strftime("%Y-%m-%d") if s.invoice_date else "",
            s.eway_bill_number or "",
            s.eway_bill_date.strftime("%Y-%m-%d") if s.eway_bill_date else "",
            s.eway_bill_expiry_date.strftime("%Y-%m-%d %H:%M") if s.eway_bill_expiry_date else "",
            s.eway_bill_status or "VEHICLE NUMBER UPDATED",
            s.new_extended_eway_bill_date.strftime("%Y-%m-%d") if s.new_extended_eway_bill_date else "",
            s.origin or "",
            s.destination or "",
            s.consignor or "",
            s.consignee or "",
            s.driver_name or "",
            s.driver_phone or "",
            s.status.value.replace("_", " ").upper(),
        ])

    output.seek(0)
    filename = f"kalebudde_shipments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/shipments/export-excel")
def export_excel_shipments(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    query = db.query(Shipment)
    if user.role == Role.CLIENT:
        query = query.filter(Shipment.client_id == user.id)
    shipments = query.order_by(Shipment.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipments"
    ws.append(HEADERS)

    for s in shipments:
        ws.append([
            s.lr_number or s.tracking_number,
            s.invoice_number or "",
            s.invoice_date.strftime("%Y-%m-%d") if s.invoice_date else "",
            s.eway_bill_number or "",
            s.eway_bill_date.strftime("%Y-%m-%d") if s.eway_bill_date else "",
            s.eway_bill_expiry_date.strftime("%Y-%m-%d %H:%M") if s.eway_bill_expiry_date else "",
            s.eway_bill_status or "VEHICLE NUMBER UPDATED",
            s.new_extended_eway_bill_date.strftime("%Y-%m-%d") if s.new_extended_eway_bill_date else "",
            s.origin or "",
            s.destination or "",
            s.consignor or "",
            s.consignee or "",
            s.driver_name or "",
            s.driver_phone or "",
            s.status.value.replace("_", " ").upper(),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"kalebudde_shipments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/shipments/sample-template")
def download_sample_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Import Template"
    ws.append(HEADERS)

    # Sample Rows for demonstration
    ws.append([
        "LR-883910",
        "INV-2026-8801",
        "2026-08-01",
        "311099214566",
        "2026-08-01",
        "2026-08-07 18:00",
        "VEHICLE NUMBER UPDATED",
        "2026-08-10",
        "Hubli, KA",
        "Bengaluru, KA",
        "Asian Paints Hubli Depot",
        "Sri Venkateshwara Traders",
        "Ramesh Kumar",
        "+91 98450 12345",
        "IN TRANSIT",
    ])
    ws.append([
        "LR-883911",
        "INV-2026-8802",
        "2026-08-02",
        "311099214567",
        "2026-08-02",
        "2026-08-08 12:00",
        "NEAR EXPIRY ALERT BEFORE 24 HR",
        "",
        "Dharwad, KA",
        "Mumbai, MH",
        "Kalebudde Warehousing",
        "Navkar Logistics Depot",
        "Suresh Patil",
        "+91 94481 67890",
        "BOOKED",
    ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kalebudde_sample_import_template.xlsx"},
    )


@router.get("/shipments/{shipment_id}", response_model=ShipmentOut)
def get_shipment(
    shipment_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)
):
    s = db.get(Shipment, shipment_id)
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")
    if user.role == Role.CLIENT and s.client_id != user.id:
        raise HTTPException(status_code=403, detail="Not your shipment")
    return s


@router.post("/shipments", response_model=ShipmentOut, status_code=201)
def create_shipment(
    payload: ShipmentCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_staff),
):
    s = Shipment(**payload.model_dump(), tracking_number=generate_tracking_number(db))
    db.add(s)
    db.flush()
    db.add(
        TrackingEvent(
            shipment_id=s.id,
            status=s.status,
            location=s.origin,
            note="Shipment booked with Kalebudde Logistics",
        )
    )
    db.commit()
    db.refresh(s)
    return s


@router.patch("/shipments/{shipment_id}", response_model=ShipmentOut)
def update_shipment(
    shipment_id: int,
    payload: ShipmentUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_staff),
):
    s = db.get(Shipment, shipment_id)
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")
    data = payload.model_dump(exclude_unset=True)
    new_status = data.get("status")
    for k, v in data.items():
        setattr(s, k, v)
    if new_status and new_status != ShipmentStatus(s.status):
        db.add(
            TrackingEvent(
                shipment_id=s.id,
                status=new_status,
                location=s.destination,
                note="Status updated",
            )
        )
    db.commit()
    db.refresh(s)
    return s


# ---------- STRICT ADMIN ONLY DELETE ----------
@router.delete("/shipments/{shipment_id}", status_code=204)
def delete_shipment(
    shipment_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)
):
    s = db.get(Shipment, shipment_id)
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")
    db.delete(s)
    db.commit()


# ---------- EXCEL & CSV BULK IMPORT ----------
@router.post("/shipments/upload-excel")
def upload_excel_csv_shipments(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_staff),
):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in [".xlsx", ".xls", ".csv"]:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file format. Please upload .xlsx, .xls or .csv file.",
        )

    rows = []
    contents = file.file.read()

    if ext == ".csv":
        text = contents.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        sheet = wb.active
        header = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                header = [str(cell).strip() if cell is not None else f"col_{idx}" for idx, cell in enumerate(row)]
            else:
                if not any(row):
                    continue
                row_dict = {}
                for idx, cell_value in enumerate(row):
                    key = header[idx] if idx < len(header) else f"col_{idx}"
                    row_dict[key] = str(cell_value).strip() if cell_value is not None else ""
                rows.append(row_dict)

    if not rows:
        raise HTTPException(status_code=400, detail="Uploaded spreadsheet file is empty.")

    created_count = 0
    for r in rows:
        def get_val(keys):
            for k in keys:
                for rk in r.keys():
                    if rk.strip().lower() == k.lower():
                        return str(r[rk]).strip()
            return ""

        lr_num = get_val(["LR NUMBER", "LR NO", "LR NUMBER / WAYBILL", "LR", "TRACKING NUMBER"])
        inv = get_val(["INVOICE", "INVOICE NUMBER", "INVOICE NO"])
        inv_date = parse_date_str(get_val(["INVOICE DATE"]))
        eway = get_val(["E-WAY BILL NUMBER", "EWAY BILL NUMBER", "E-WAY BILL NO", "EWAY BILL NO"])
        eway_date = parse_date_str(get_val(["E-WAY BILL DATE", "EWAY BILL DATE"]))
        eway_exp = parse_date_str(get_val(["E-WAY BILL EXPIRY DATE", "EWAY BILL EXPIRY DATE"]))
        eway_status = get_val(["E-WAY BILL STATUS", "EWAY BILL STATUS"]) or "VEHICLE NUMBER UPDATED"
        new_ext_date = parse_date_str(get_val(["NEW EXTENDED E-WAY BILL DATE", "EXTENDED E-WAY BILL DATE"]))
        origin = get_val(["ORIGIN", "FROM"]) or "Hubli, KA"
        dest = get_val(["DESTINATION", "TO"]) or "Bengaluru, KA"
        consignor = get_val(["CONSIGNOR", "SENDER"]) or "Kalebudde Logistics"
        consignee = get_val(["CONSIGNEE", "RECEIVER"]) or "Valued Client"
        driver_name = get_val(["DRIVER NAME", "DRIVER"])
        driver_no = get_val(["DRIVER NO", "DRIVER PHONE", "DRIVER MOBILE"])
        delivery_status_raw = get_val(["DELIVERY STATUS", "STATUS"]).lower().replace(" ", "_")

        status_enum = ShipmentStatus.BOOKED
        for st in ShipmentStatus:
            if st.value == delivery_status_raw:
                status_enum = st
                break

        tracking_code = lr_num or generate_tracking_number(db)

        shipment = Shipment(
            tracking_number=tracking_code,
            lr_number=lr_num or tracking_code,
            invoice_number=inv or f"INV-{uuid.uuid4().hex[:6].upper()}",
            invoice_date=inv_date or datetime.now(timezone.utc),
            eway_bill_number=eway or f"EWB{random.randint(100000000000, 999999999999)}",
            eway_bill_date=eway_date or datetime.now(timezone.utc),
            eway_bill_expiry_date=eway_exp or datetime.now(timezone.utc),
            eway_bill_status=eway_status,
            new_extended_eway_bill_date=new_ext_date,
            origin=origin,
            destination=dest,
            consignor=consignor,
            consignee=consignee,
            driver_name=driver_name,
            driver_phone=driver_no,
            status=status_enum,
        )
        db.add(shipment)
        created_count += 1

    db.commit()
    return {"message": f"Successfully imported {created_count} shipments", "imported_count": created_count}


@router.post(
    "/shipments/{shipment_id}/events", response_model=TrackingEventOut, status_code=201
)
def add_event(
    shipment_id: int,
    payload: TrackingEventCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_staff),
):
    s = db.get(Shipment, shipment_id)
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")
    ev = TrackingEvent(
        shipment_id=s.id,
        status=payload.status,
        location=payload.location,
        note=payload.note,
        occurred_at=payload.occurred_at or datetime.now(timezone.utc),
    )
    s.status = payload.status
    db.add(ev)
    db.commit()
    db.refresh(ev)
    return ev
